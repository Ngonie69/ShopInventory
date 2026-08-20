"""Regenerate ShopInventory.Web/Common/DeliveryRoutes.g.cs from the routes workbook.

The workbook lists each delivery route as a pair of columns -- a business partner
code and a shop name -- under a day-of-week heading. Neither column can be
trusted on its own: 28 of the codes it carries do not exist in SAP at all, and
one (tmp013, "TM CHIREDZI") names a shop in the wrong province. So every stop is
resolved against a live SAP business-partner dump and only the codes SAP
confirms are written out.

Usage:
    python generate_delivery_routes.py --workbook "REVISED  ROUTES 2026.xlsx" \
        --session <B1SESSION id> [--service-layer https://10.10.10.6:50000/b1s/v1/]
    python generate_delivery_routes.py --workbook ... --customers customers.json

--session fetches the partner master itself; --customers reuses a saved dump so
the generator can be re-run without a live session (pass --save-customers to
write one). Re-running with the same inputs reproduces the same file.
"""

from __future__ import annotations

import argparse
import collections
import json
import re
import ssl
import sys
import unicodedata
import urllib.parse
import urllib.request
from pathlib import Path

SHEET = "PROPOSED ROUTES"
DAYS = ("MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY")

# (code column, name column) pairs. The sheet lays routes out side by side.
COLUMN_PAIRS = ((2, 3), (4, 5), (6, 7), (8, 9), (10, 11), (12, 13), (14, 15))

# Markers the sheet writes into the code column instead of a partner code.
# "NOT ON THIS ROUTE" and "CLOSED" retire the stop; "FACTORY DIRECT" only says
# the drop is not made off this truck, so the shop stays on the route.
DROP_MARKERS = {"NOT ON THIS ROUTE", "CLOSED"}

# Rows naming a group of shops rather than one, so they resolve to nothing.
NOT_A_PARTNER = {
    "NORTH& WEST ROUTES SHOPS",
    "PNP SOUTH & B/DALE ROUTES SHOPS",
    "MSASA &PNP CENTRAL SHOPS",
    "LAN107",
}

# Stops the sheet leaves without a usable code, resolved against SAP by hand and
# recorded here so the run stays reproducible. Values are SAP codes minus any
# currency suffix -- every currency variant of the partner is picked up from them.
OVERRIDES = {
    "AMP WESTGATE": ["ASS005", "ASS010"],
    "AMP COVENTRY": ["ASS006", "ASS009", "ASS012"],
    "BHOLA BISHOPGAUL": ["BHO015"],
    "BHOLA AVONLEA": ["BHO025"],
    "BHOLA CHINHOYI": ["BHO010"],
    "BHOLA MANDELA": ["BHO020"],
    "BHOLA MUGABE": ["BHO018"],
    "BHOLA TYWALD": ["BHO011"],
    "BHOLA TYNWALD": ["BHO011"],
    "BHOLA W/PARK": ["BHO017"],
    "BHOLA W/CLIFF": ["BHO012"],
    "BHOLA WORKINGTON": ["BHO014"],
    "BHOLA HOGERTY": ["BHO006"],
    # Mabvuku carries a Mega Mart and a Supermarket; the sheet lists one stop.
    "BHOLA MABVUKU": ["BHO019", "BHO021"],
    "MEGASAVE CHINHOYI": ["LAN017"],
    "MEGA SAVE GLENDALE": ["LAN012"],
    "SAV MARONDERA": ["SAV001", "SAV002"],
    "GAINS GRANITESIDE": ["GAI006"],
    "TM K KAUNDA": ["TMP011"],
    "CREDLE ZVISHAVANE": ["CRE006"],
    # The sheet codes this tmp013, which SAP holds as TM Budiriro in Harare.
    "TM CHIREDZI": ["TMP039"],
    "CHIRUNDU CANTEEN": ["CHI010"],
    "SAI MART GWERU": ["SAI014"],
    "PANMART MUTARE": ["MAU002"],
    "PANMART RUSAPE": ["MAU001"],
    "SPAR MAZOWE": ["SPA076"],
    "NYANINGWE W/PARK": ["NYA009"],
    "ZUVA AERODROME": ["CYB004"],
    "PAYE MARKETING": ["PAY002"],
    "NR NYIKA": ["NRI060"],
}

# Shops the workbook omits entirely, added on top of it. Each one invoices
# regularly and sits in a town the route demonstrably already stops in, so the
# route is evidenced rather than guessed -- the justifying stop is named beside
# it. Keep that discipline when adding: if no existing stop on the route shares
# the town, it belongs in the open questions below, not here.
#
# Deliberately NOT added: the Bulawayo and Matabeleland shops (TM Lobengula,
# PnP Bradfield/Ascot/Fife Street/Gwanda, Greens, Fresh and Green, Fazak, the
# Sai Mart estate, Metro Gwanda). No route stops in that region -- the workbook's
# BULAWAYO route is a single 24T run to CHEESEMAN BULAWAYO, the depot, which then
# distributes locally. Putting those shops on it would claim a Harare truck calls
# at each one.
ADDITIONS = {
    "MARONDERA-CHIPINGE": [
        # The route is named for Marondera and runs through it, but the workbook
        # gives it no Marondera stop at all.
        "BHO023",   # BHOLA SUPERMARKET - MARONDERA
        "SPA067",   # SPAR Marondera
        "TMP089",   # TM Main Street Marondera
        "NRI034",   # N Richards Marondera
        "LAN013",   # Megasave Marondera
        # Mutare and Chipinge: the route already stops at SPAR Mutare, TM Main
        # Street Mutare, TM Sakubva and TM Chipinge.
        "BHO013",   # Bhola Supermarket Mutare Town
        "GAI102",   # Metropeech Hypermarket Mutare
        "GAI113",   # Gains Cash and Carry Sakubva (Mutare)
        "GAI118",   # Gain Cash & Carry Wholesale Checheche (Chipinge district)
    ],
    "MIDLANDS 1": [
        "GAI114",   # Metro Peech Hypermarket Kwekwe -- route stops at SPAR/TM/NR Kwekwe
        "SPA075",   # Spar Express Kadoma -- route stops at TM, SPAR and OK Kadoma
        "GAI078",   # Metro Hyper Gweru -- route stops at PNP, Sai Mart, OK and NR Gweru
    ],
    "MIDLANDS 2": [
        "GAI054",   # GAINS CHIREDZI -- route stops at TM Chiredzi and NR Chiredzi
        "GAI111",   # Metro Hypermarket Chiredzi
        "GAI110",   # Gains Cash and Carry Triangle -- route stops at TM Triangle
    ],
}

CURRENCY_SUFFIX = re.compile(r"\b(USD|US\$|FCA|ZIG|RTGS|BOND)\b", re.IGNORECASE)
BP_CODE = re.compile(r"^[A-Za-z]{3}\d{3}$")
LEADING_CODE = re.compile(r"^\s*([A-Za-z]{2,4}\s?\d{2,4})")

ALIASES = {
    "PNP": "PICK N PAY", "NR": "N RICHARDS", "FW": "FOOD WORLD",
    "MEGA": "MEGASAVE", "GAINS": "GAIN", "CM": "CHEESEMAN",
    "CC": "CASH AND CARRY", "WSALE": "WHOLESALE", "MTRE": "MUTARE",
}
NOISE = {"THE", "AND", "OF", "PVT", "LTD", "TA", "USD", "FCA",
         "ZIG", "RTGS", "PL", "P", "L", "ENT", "INV"}


def ascii_fold(value: str) -> str:
    return unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()


def name_key(value: str) -> str:
    """Collapse a partner name to the shop it names, dropping the currency tag."""
    folded = CURRENCY_SUFFIX.sub(" ", ascii_fold(value).upper())
    folded = folded.replace("&", " AND ").replace("/", " ").replace("-", " ")
    folded = folded.replace("'", "").replace(".", " ")
    folded = re.sub(r"\(.*?\)", " ", folded)
    return re.sub(r"\s+", " ", folded).strip()


def base_code(card_code: str) -> str:
    match = LEADING_CODE.match(ascii_fold(card_code))
    if not match:
        return ascii_fold(card_code).strip().upper()
    return re.sub(r"\s+", "", match.group(1)).upper()


def tokens(value: str) -> set:
    out = []
    for token in name_key(value).split():
        out.extend(part for part in ALIASES.get(token, token).split() if part not in NOISE)
    return set(out)


def fetch_customers(service_layer: str, session_id: str) -> list:
    context = ssl.create_default_context()
    context.check_hostname = False
    context.verify_mode = ssl.CERT_NONE
    query = urllib.parse.quote(
        "$select=CardCode,CardName,CardType&$filter=CardType eq 'cCustomer'&$orderby=CardCode",
        safe="$=&,'()",
    )
    rows = []
    skip = 0
    while True:
        request = urllib.request.Request(
            f"{service_layer.rstrip('/')}/BusinessPartners?{query}&$skip={skip}",
            headers={
                "Cookie": f"B1SESSION={session_id}",
                # Service Layer answers 20 rows without this, whatever $top says.
                "Prefer": "odata.maxpagesize=500",
                "Accept": "application/json",
            },
        )
        with urllib.request.urlopen(request, context=context, timeout=180) as response:
            page = json.loads(response.read().decode("utf-8")).get("value", [])
        rows.extend(page)
        if len(page) < 500:
            return rows
        skip += len(page)


def read_stops(workbook_path: Path) -> list:
    import openpyxl

    sheet = openpyxl.load_workbook(workbook_path, data_only=True)[SHEET]

    def cell(row: int, column: int) -> str:
        value = sheet.cell(row=row, column=column).value
        return "" if value is None else str(value).strip()

    day_rows = [(row, cell(row, 2).upper())
                for row in range(1, sheet.max_row + 1)
                if cell(row, 2).upper() in DAYS]
    day_rows.append((sheet.max_row + 1, ""))

    routes = []
    for index in range(len(day_rows) - 1):
        header_row, day = day_rows[index]
        if not day:
            break
        last_row = day_rows[index + 1][0] - 1
        for code_column, name_column in COLUMN_PAIRS:
            route = cell(header_row, name_column)
            if not route:
                continue
            stops = []
            for row in range(header_row + 3, last_row + 1):
                name = cell(row, name_column)
                code = cell(row, code_column)
                if not name or name.upper().startswith("TRUCK ALLOC"):
                    continue
                if code.upper() in DROP_MARKERS or name.upper() in NOT_A_PARTNER:
                    continue
                stops.append({"name": name, "code": code if BP_CODE.match(code) else ""})
            if stops:
                routes.append({
                    "day": day.capitalize(),
                    "route": route,
                    "truck": cell(header_row + 2, name_column),
                    "stops": stops,
                })
    return routes


def canonical_route(route: str) -> str:
    """Fold the sheet's overflow columns back into the route they continue."""
    trimmed = re.sub(r"\s*-?\s*CONTINUED\s*$", "", route.strip(), flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", trimmed).strip()


def build(workbook_path: Path, customers: list):
    # A saved dump may hold every partner type. Suppliers share names with
    # customers -- "Cheese Galore ( Packaging)" is both -- so leaving them in
    # silently adds supplier codes to a delivery route. Filter rather than trust
    # the caller. Rows with no CardType are kept: --session already filtered.
    customers = [c for c in customers
                 if (c.get("CardType") or "cCustomer") == "cCustomer"]
    for customer in customers:
        customer["CardName"] = (customer.get("CardName") or "").strip()

    by_name = collections.defaultdict(list)
    for customer in customers:
        by_name[name_key(customer["CardName"])].append(customer)
    by_base = {}
    for customer in customers:
        by_base.setdefault(base_code(customer["CardCode"]), customer)

    def variants_of(codes):
        """Every currency variant of the shops these codes name."""
        found = {}
        for code in codes:
            customer = by_base.get(base_code(code))
            if customer is None:
                continue
            for variant in by_name[name_key(customer["CardName"])]:
                found[variant["CardCode"]] = variant
        return list(found.values())

    def by_exact_name(name):
        query = tokens(name)
        if not query:
            return []
        for key, members in by_name.items():
            candidate = tokens(key)
            if candidate and query <= candidate:
                return members
        return []

    merged = {}
    unresolved = []
    for route in read_stops(workbook_path):
        key = canonical_route(route["route"])
        entry = merged.setdefault(key, {"route": key, "days": [], "trucks": [], "codes": {}})
        if route["day"] not in entry["days"]:
            entry["days"].append(route["day"])
        if route["truck"] and route["truck"] not in entry["trucks"]:
            entry["trucks"].append(route["truck"])
        for stop in route["stops"]:
            if stop["name"].upper() in OVERRIDES:
                matches = variants_of(OVERRIDES[stop["name"].upper()])
            elif stop["code"]:
                matches = variants_of([stop["code"]])
            else:
                matches = []
            if not matches:
                matches = by_exact_name(stop["name"])
            if not matches:
                unresolved.append({"route": key, "day": route["day"], "name": stop["name"]})
                continue
            for match in matches:
                entry["codes"][match["CardCode"]] = match["CardName"]

    # Shops the workbook leaves off their route entirely. Applied after the sheet
    # so a stop the sheet does list is never overwritten by one of these.
    for route, codes in ADDITIONS.items():
        entry = merged.get(route)
        if entry is None:
            raise SystemExit(f"ADDITIONS names route {route!r}, which the workbook does not define")
        for code in codes:
            matches = variants_of([code])
            if not matches:
                raise SystemExit(f"ADDITIONS code {code!r} for {route!r} matches no SAP customer")
            for match in matches:
                entry["codes"][match["CardCode"]] = match["CardName"]

    catalogue = [entry for entry in merged.values() if entry["codes"]]
    catalogue.sort(key=lambda entry: entry["route"])
    return catalogue, unresolved


def emit(catalogue: list, workbook_name: str) -> str:
    lines = [
        "// <auto-generated>",
        "//     Generated by scripts/DeliveryRoutes/generate_delivery_routes.py from",
        f'//     "{workbook_name}", sheet "{SHEET}", with every business partner code',
        "//     resolved against the SAP customer master. Do not edit by hand -- re-run",
        "//     the generator when the routes change.",
        "// </auto-generated>",
        "",
        "namespace ShopInventory.Web.Common;",
        "",
        "public static partial class DeliveryRoutes",
        "{",
        "    private static readonly DeliveryRoute[] RouteTable =",
        "    [",
    ]
    for entry in catalogue:
        days = ", ".join(f'"{day}"' for day in entry["days"])
        trucks = ", ".join(f'"{truck}"' for truck in entry["trucks"])
        lines.append(f'        new("{entry["route"]}", [{days}], [{trucks}],')
        lines.append("        [")
        for code, name in sorted(entry["codes"].items()):
            comment = re.sub(r"\s+", " ", ascii_fold(name)).strip()
            lines.append(f'            "{code}", // {comment}')
        lines.append("        ]),")
    lines += ["    ];", "}", ""]
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--session")
    parser.add_argument("--service-layer", default="https://10.10.10.6:50000/b1s/v1/")
    parser.add_argument("--customers", type=Path)
    parser.add_argument("--save-customers", type=Path)
    parser.add_argument("--output", type=Path,
                        default=Path("ShopInventory.Web/Common/DeliveryRoutes.g.cs"))
    args = parser.parse_args()

    if args.customers and args.customers.exists():
        customers = json.loads(args.customers.read_text(encoding="utf-8"))
    elif args.session:
        customers = fetch_customers(args.service_layer, args.session)
    else:
        parser.error("pass --session to fetch the partner master, or --customers to reuse a dump")

    if args.save_customers:
        args.save_customers.write_text(json.dumps(customers, indent=0), encoding="utf-8")

    catalogue, unresolved = build(args.workbook, customers)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(emit(catalogue, args.workbook.name), encoding="utf-8")

    codes = sum(len(entry["codes"]) for entry in catalogue)
    print(f"{args.output}: {len(catalogue)} routes, {codes} business partner codes")
    if unresolved:
        print(f"\n{len(unresolved)} stops have no SAP customer and are not in the catalogue:")
        for stop in unresolved:
            print(f"    {stop['route']:<28} {stop['name']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
